# Author: Authey
# Date: 01/03/2021
import os

from PIL import Image
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def rgb_to_hex(rgb):
    rb = list(rgb)
    st = ''
    for cl in rb:
        st += hex(cl)[2:].rjust(2, '0')
    if st.count('0') == 6:
        st = st[:-1] + '1'
    return st


def ch26(x):
    chr26 = [chr(65 + i) for i in range(26)]
    zh = ''
    while x != -1:
        zh = chr26[x % 26] + zh
        x = x // 26 - 1
    return zh


class Converter:

    def __init__(self, pic, xls):
        self.pic = pic
        self.xls = xls
        self.data = list()
        self.x = 0
        self.y = 0

    def get_pix(self):
        img_src = Image.open(self.pic)
        self.x = img_src.size[0]
        self.y = img_src.size[1]
        img_src = img_src.convert('RGB')
        str_list = img_src.load()
        for row in range(self.y):
            self.data.append(list())
            for col in range(self.x):
                self.data[row].append(str_list[col, row])
        img_src.close()

    def fill_cells(self):
        if not os.path.isfile(self.xls):
            wb = Workbook()
            ws = wb.active
            ws.title = "pic"
            wb.save(self.xls)
        wb = load_workbook(self.xls)
        ws = wb.active
        for row in range(self.y):
            ws.row_dimensions[row+1].height = 15
            for col in range(self.x):
                ws.column_dimensions[ch26(col)].width = 2.78
                c_rgb = self.data[row][col]
                c_hex = rgb_to_hex(c_rgb)
                fills = PatternFill('solid', fgColor=c_hex)
                ws[f'{ch26(col)}{row+1}'].fill = fills
        wb.save(self.xls)
        wb.close()


if __name__ == '__main__':
    if len(sys.argv) == 3:
        pic_path = sys.argv[1]
        xls_path = sys.argv[2]
        pic_xls = Converter(pic_path, xls_path)
        pic_xls.get_pix()
        pic_xls.fill_cells()
        print('Done!')
    else:
        print('Invalid number of arguments')
        exit()
